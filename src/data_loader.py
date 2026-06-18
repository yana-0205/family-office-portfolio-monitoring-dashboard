from __future__ import annotations

from dataclasses import dataclass
from difflib import get_close_matches
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.config import CSV_DIR, DOCUMENTS_DIR, WORKBOOK_PATH


@dataclass(frozen=True)
class TableReference:
    source: str
    name: str
    location: Path | None
    exact_match: bool
    matched_alias: str


def _normalize_name(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def list_csv_files() -> list[Path]:
    return sorted(CSV_DIR.glob("*.csv"))


def list_pdf_files() -> list[Path]:
    return sorted(DOCUMENTS_DIR.glob("*.pdf"))


def list_excel_sheets() -> list[str]:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    workbook = load_workbook(WORKBOOK_PATH, read_only=True)
    return list(workbook.sheetnames)


def _csv_lookup() -> dict[str, Path]:
    lookup: dict[str, Path] = {}
    for path in list_csv_files():
        lookup[_normalize_name(path.stem)] = path
        lookup[_normalize_name(path.name)] = path
    return lookup


def _sheet_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for sheet_name in list_excel_sheets():
        lookup[_normalize_name(sheet_name)] = sheet_name
    return lookup


def safe_find_csv(possible_names: list[str] | tuple[str, ...]) -> Path | None:
    aliases = [name for name in possible_names if name]
    csv_lookup = _csv_lookup()

    for alias in aliases:
        normalized = _normalize_name(alias)
        if normalized in csv_lookup:
            return csv_lookup[normalized]

    csv_candidates = sorted({path.stem for path in list_csv_files()})
    normalized_csv = {_normalize_name(name): name for name in csv_candidates}
    for alias in aliases:
        match = get_close_matches(
            _normalize_name(alias),
            list(normalized_csv.keys()),
            n=1,
            cutoff=0.75,
        )
        if match:
            matched_name = normalized_csv[match[0]]
            return CSV_DIR / f"{matched_name}.csv"

    return None


def safe_find_table(possible_names: list[str] | tuple[str, ...]) -> TableReference | None:
    aliases = [name for name in possible_names if name]
    csv_path = safe_find_csv(aliases)
    if csv_path is not None:
        matched_alias = next(
            (
                alias
                for alias in aliases
                if _normalize_name(alias) in (_normalize_name(csv_path.stem), _normalize_name(csv_path.name))
            ),
            aliases[0],
        )
        exact_match = any(
            _normalize_name(alias) in (_normalize_name(csv_path.stem), _normalize_name(csv_path.name))
            for alias in aliases
        )
        return TableReference(
            source="csv",
            name=csv_path.stem,
            location=csv_path,
            exact_match=exact_match,
            matched_alias=matched_alias,
        )

    sheet_lookup = _sheet_lookup()
    for alias in aliases:
        normalized = _normalize_name(alias)
        if normalized in sheet_lookup:
            sheet_name = sheet_lookup[normalized]
            return TableReference(
                source="excel",
                name=sheet_name,
                location=WORKBOOK_PATH,
                exact_match=True,
                matched_alias=alias,
            )

    csv_candidates = sorted({path.stem for path in list_csv_files()})
    normalized_csv = {_normalize_name(name): name for name in csv_candidates}
    for alias in aliases:
        match = get_close_matches(
            _normalize_name(alias),
            list(normalized_csv.keys()),
            n=1,
            cutoff=0.75,
        )
        if match:
            matched_name = normalized_csv[match[0]]
            path = CSV_DIR / f"{matched_name}.csv"
            return TableReference(
                source="csv",
                name=matched_name,
                location=path,
                exact_match=False,
                matched_alias=alias,
            )

    excel_candidates = list_excel_sheets()
    normalized_sheets = {_normalize_name(name): name for name in excel_candidates}
    for alias in aliases:
        match = get_close_matches(
            _normalize_name(alias),
            list(normalized_sheets.keys()),
            n=1,
            cutoff=0.75,
        )
        if match:
            matched_name = normalized_sheets[match[0]]
            return TableReference(
                source="excel",
                name=matched_name,
                location=WORKBOOK_PATH,
                exact_match=False,
                matched_alias=alias,
            )

    return None


def read_csv_table(table_name_or_filename: str) -> pd.DataFrame:
    reference = safe_find_table([table_name_or_filename])
    if reference is None or reference.source != "csv" or reference.location is None:
        raise FileNotFoundError(
            f"CSV table not found for '{table_name_or_filename}' in {CSV_DIR}"
        )
    return pd.read_csv(reference.location)


def read_excel_sheet(sheet_name: str) -> pd.DataFrame:
    reference = safe_find_table([sheet_name])
    if reference is None or reference.source != "excel":
        raise FileNotFoundError(
            f"Excel sheet not found for '{sheet_name}' in {WORKBOOK_PATH}"
        )
    return pd.read_excel(WORKBOOK_PATH, sheet_name=reference.name)


def load_required_tables() -> dict[str, pd.DataFrame]:
    required_table_aliases = {
        "portfolio_monthly_summary": [
            "portfolio_monthly_summary",
            "portfolio monthly summary",
        ],
        "portfolio_holdings": ["portfolio_holdings", "portfolio holdings"],
        "private_fund_positions": [
            "private_fund_positions",
            "private positions pre ingestion",
        ],
        "cash_accounts": ["cash_accounts", "cash accounts"],
        "document_metadata": ["document_metadata", "document metadata"],
        "ground_truth_extractions": [
            "ground_truth_extractions",
            "ground truth extractions",
        ],
        "validation_rules": ["validation_rules", "validation rules"],
        "table_name_map": ["table_name_map", "table name map"],
    }

    tables: dict[str, pd.DataFrame] = {}
    missing: list[str] = []
    for logical_name, aliases in required_table_aliases.items():
        reference = safe_find_table(aliases)
        if reference is None:
            missing.append(logical_name)
            continue
        if reference.source == "csv":
            tables[logical_name] = pd.read_csv(reference.location)
        else:
            tables[logical_name] = pd.read_excel(WORKBOOK_PATH, sheet_name=reference.name)

    if missing:
        raise FileNotFoundError(
            "Required tables could not be located: " + ", ".join(sorted(missing))
        )

    return tables
